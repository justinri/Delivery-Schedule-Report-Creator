########## Release Notes:
import openpyxl, xlsxwriter
from pathlib import Path
import datetime


#******************************** Only Edit **************************/
### Enter filename here:
filename = "Job_Schedule"
out_How_Many_Weeks = 5

#******************************** STOP!!!!! **************************/
#*****************************************************************/
wb_obj = openpyxl.load_workbook(filename + ".xlsx") 
sheet = wb_obj.active

### Deleting the last row "* Indicates a Job associated with a Sales Order."
sheet.delete_rows(sheet.max_row)

# Create a workbook and add a worksheet.
workbook = xlsxwriter.Workbook(filename + "_output" + ".xlsx")
worksheet = { "Jobs" : workbook.add_worksheet("Jobs"), 
			  "Error- Due Before 2020" : workbook.add_worksheet("Error- Due Before 2020"),
			  "Error- Remaing quantities are 0" : workbook.add_worksheet("Error- Remaing quantities are 0"),
			  "Error- Job Remaing less than 0" : workbook.add_worksheet("Error- Job Remaing less than 0")}

whatSheet = "Jobs"
row  = { "Jobs" : 1, "Error- Due Before 2020" : 1, "Error- Remaing quantities are 0" : 1, "Error- Job Remaing less than 0": 1}

today = datetime.date.today()
mondays = [today + datetime.timedelta(days=-today.weekday(), weeks=num) for num in range(out_How_Many_Weeks)]

### Looping through Jobs
for num, rows in enumerate(sheet.iter_rows()):
	if num == 0: continue    ### Skpping the header row

	### If the year is before 2020, it probably a practice Job and is an error. 
	tmpDate = sheet.cell(row=num+1, column=8).value.date()
	year = int(str(tmpDate)[0:4])

	### If sales orders and Job quantities are both zero, it is probably an error
	orderQTY = sheet.cell(row=num+1, column=4).value
	completedQty = sheet.cell(row=num+1, column=5).value
	soQTY  = sheet.cell(row=num+1, column=7).value ### Remaining on SO
	jobQTY = orderQTY - completedQty 			   ### Remaining on Job
	if year < 2020:
		whatSheet = "Error- Due Before 2020"
	elif soQTY == 0 and jobQTY == 0:
		whatSheet = "Error- Remaing quantities are 0"
	elif jobQTY < 0:
		whatSheet = "Error- Job Remaing less than 0"		
	else:
		whatSheet = "Jobs"

	### Putting the year last
	dueDate = str(tmpDate)[5:] + "-" +str(year)

	worksheet[whatSheet].write(row[whatSheet], 0, sheet.cell(row=num+1, column=1).value) ### Job Number
	worksheet[whatSheet].write(row[whatSheet], 1, sheet.cell(row=num+1, column=2).value) ### Part Number
	worksheet[whatSheet].write(row[whatSheet], 2, sheet.cell(row=num+1, column=3).value) ### Description
	worksheet[whatSheet].write(row[whatSheet], 3, orderQTY) 							 ### Order Qty
	worksheet[whatSheet].write(row[whatSheet], 4, completedQty) 						 ### Completed Qty
	worksheet[whatSheet].write(row[whatSheet], out_How_Many_Weeks+7, dueDate) 			 ### Actual Due Date
	
	### Creating the due dates
	shipDate = sheet.cell(row=num+1, column=8).value.date()
	print(shipDate)
	if shipDate < mondays[0]: tmpCol = 5 ### Then date is past due
	elif shipDate > mondays[-1]: tmpCol = out_How_Many_Weeks+6 ### Then date is a future order
	else:
		whatMonday = shipDate + datetime.timedelta(days=-shipDate.weekday(), weeks=0)
		for num3, tmpMonday in enumerate(mondays):
			if whatMonday == tmpMonday:
				tmpCol = 6+num3
	worksheet[whatSheet].write(row[whatSheet], tmpCol, str(soQTY) + " (" + str(jobQTY)+ ")")
	row[whatSheet] += 1

### Adding headers and changing cells widths
bold = workbook.add_format({'bold': True, 'num_format': 'd mmm yyyy'})
headers = ["Job", "Part Number", "Description", "Order Qty", "Completed Qty", "PAST DUE"]
[headers.append(items) for items in mondays]
headers.append("FUTURE")
headers.append("Actual Due Date")
for key in worksheet:
	### Heading headers
	for num, item in enumerate(headers):
		worksheet[key].write(0,num, item, bold)

	### Changing cells widths
	worksheet[key].set_column(0,0, 6)
	worksheet[key].set_column(1,1, 11)
	worksheet[key].set_column(2,2, 16)
	worksheet[key].set_column(3,3, 12)
	worksheet[key].set_column(4,4, 12)
	worksheet[key].set_column(5,5, 12)
	worksheet[key].set_column(6,out_How_Many_Weeks+7, 11)
	worksheet[key].set_column(out_How_Many_Weeks+8,out_How_Many_Weeks+8, 15)
	worksheet[key].freeze_panes(1, 0)
workbook.close()
