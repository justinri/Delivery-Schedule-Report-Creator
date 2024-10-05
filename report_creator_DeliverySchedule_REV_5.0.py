########## Release Notes:
# Includes Daimler Check and added more customer numbers for John Deere

import openpyxl, xlsxwriter
from pathlib import Path
import datetime, sys
sys.path.append('daimler_check')
from daimler_check_rev_1 import checkDaimler


#******************************** Only Edit **************************/
### Enter filename here:
filename = "SH_DeliverySchedule"
filenameDaimler = "862"
stockFilename = "MC_StockStatus"
out_How_Many_Weeks = 5
includeDaimlerCheck = "yes"

#******************************** STOP!!!!! **************************/
### Finding on hand amounts
onHand = {}
wb_obj = openpyxl.load_workbook(stockFilename + ".xlsx") 
sheet = wb_obj.active

for num, rows in enumerate(sheet.iter_rows()):
	if num == 0 or num == 1: continue    ### Skpping the header rows
	for num2, cell in enumerate(rows):
		if num2 != 0: break
		try:
			qty_onHand_units = str(sheet.cell(row=num+1, column=3).value) + " " + sheet.cell(row=num+1, column=2).value
			onHand.update({cell.value : qty_onHand_units})
		except TypeError:
			pass

#*****************************************************************/
#### Including the Daimler Check
if includeDaimlerCheck.lower() == "yes":
	daimler862Info = checkDaimler(out_How_Many_Weeks, filenameDaimler)

#*****************************************************************/
### Creating a master list of planner codes to check against
Planner_Codes = ["CN","CP","DC","HC","HR","HV","LD","LK","LP","M","MF","PP","PR","TB","TC","TR","V"]
master_PCList = {tmp:[] for tmp in Planner_Codes}

wb_obj = openpyxl.load_workbook("MasterPartList.xlsx") 
sheet = wb_obj.active

for num, rows in enumerate(sheet.iter_rows()):
	if num == 0: continue    ### Skpping the header row
	for num2, cell in enumerate(rows):
		if cell.value != None:
			master_PCList[Planner_Codes[num2]].append(cell.value)

#*****************************************************************/
wb_obj = openpyxl.load_workbook(filename + ".xlsx") 
sheet = wb_obj.active

### Deleting the last row "* Indicates a Job associated with a Sales Order."
sheet.delete_rows(sheet.max_row)

# Create a workbook and add a worksheet.
workbook = xlsxwriter.Workbook(filename + "_output" + ".xlsx")
worksheet = { "MPS" : workbook.add_worksheet("MPS"), 
			  "Daimler" : workbook.add_worksheet("Daimler"),
			  "Daimler Aftermarket" : workbook.add_worksheet("Daimler Aftermarket"),
			  "Navistar Production" : workbook.add_worksheet("Navistar Production"),
			  "Navistar Service" : workbook.add_worksheet("Navistar Service"),
			  "AEES" : workbook.add_worksheet("AEES"),
			  "John Deere" : workbook.add_worksheet("John Deere"),
			  "BPC" : workbook.add_worksheet("BPC"),
			  "Jobs" : workbook.add_worksheet("Jobs"),
			  "Jobs- Associated with SO" : workbook.add_worksheet("Jobs- Associated with SO")}
whatSheet = "MPS"
row  = { "MPS" : 1, "Daimler" : 1, "Daimler Aftermarket" : 1, "Navistar Production" : 1, "Navistar Service" : 1, "AEES" : 1, "John Deere" : 1, "BPC" : 1, "Jobs" : 1, "Jobs- Associated with SO" : 1}
col  = { "MPS" : 0, "Daimler" : 0, "Daimler Aftermarket" : 0, "Navistar Production" : 0, "Navistar Service" : 0, "AEES" : 0, "John Deere" : 0, "BPC" : 0, "Jobs" : 0, "Jobs- Associated with SO" : 0}
tmpDaimler = {}

today = datetime.date.today()
mondays = [today + datetime.timedelta(days=-today.weekday(), weeks=num) for num in range(out_How_Many_Weeks)]

def writeToFile(plannerCode, SO, partNum, PO, onHandQty, custNum, orderStatus, description, qty, shipDate):
	""" Write all information to an .xlsx file"""
	### Finding and wiring dates to file
	if shipDate < mondays[0]: tmpCol = 6 ### Then date is past due
	elif shipDate > mondays[-1]: tmpCol = out_How_Many_Weeks+7 ### Then date is a future order
	else:
		whatMonday = shipDate + datetime.timedelta(days=-shipDate.weekday(), weeks=0)
		for num3, tmpMonday in enumerate(mondays):
			if whatMonday == tmpMonday:
				tmpCol = 7+num3

	### Wiring to file
	worksheet[whatSheet].write(row[whatSheet], 1, plannerCode)
	worksheet[whatSheet].write(row[whatSheet], 2, SO)	    ### Sales order 
	worksheet[whatSheet].write(row[whatSheet], 3, partNum)	    ### Our Part Number 
	worksheet[whatSheet].write(row[whatSheet], 4, PO)   ### POs
	worksheet[whatSheet].write(row[whatSheet], 5, onHandQty)

	worksheet[whatSheet].write(row[whatSheet], 0, custNum) 
	worksheet[whatSheet].write(row[whatSheet], out_How_Many_Weeks+8, orderStatus) 
	worksheet[whatSheet].write(row[whatSheet], out_How_Many_Weeks+9, description)  
	worksheet[whatSheet].write(row[whatSheet], tmpCol, qty)
	return 0

for num, rows in enumerate(sheet.iter_rows()):
	if num == 0: continue    ### Skpping the header row
	elif sheet.cell(row=num+1, column=1).value == None: break	

	### Does not work for some reason, unless you delete it first... but then you can undo?!?!?!?!
	### Skipping the last line of excel, it says "* Indicates a Job associated with a Sales Order."
#	if sheet.cell(row=num+1, column=1).value == "* Indicates a Job associated with a Sales Order.": continue
#	if sheet.cell(row=num+1, column=1).value[0] == "*": break

	### If it is a sales order
	if "-" in sheet.cell(row=num+1, column=1).value:
		### Checking to see what company it is
		custNum = sheet.cell(row=num+1, column=8).value
		if custNum == "2768A" or custNum == "2768B": whatSheet = "Daimler"
		elif custNum == "2768S": whatSheet = "Daimler Aftermarket"
		elif custNum == "5916A" or custNum == "5916L" or custNum == "5916LT" or custNum == "5916U" or custNum == "5934": whatSheet = "Navistar Production"
		elif custNum == "5916C"  or custNum == "5916B" or custNum == "5912": whatSheet = "Navistar Service"
		elif custNum == "0160": whatSheet = "AEES"
		elif custNum == "1880" or custNum == "1881" or custNum == "1878" or custNum == "1870" or custNum == "1876" or custNum == "1879" or custNum == "100148" or custNum == "1874" or custNum == "1877" or custNum == "1875": whatSheet = "John Deere"
		elif custNum == "0845": whatSheet = "BPC"
		else: whatSheet = "MPS"

	### Else it is a job
	elif "*" in sheet.cell(row=num+1, column=1).value: whatSheet = "Jobs- Associated with SO"
	else: whatSheet = "Jobs"

	### Creating the required variables
	SO = sheet.cell(row=num+1, column=1).value # Sales order/Job Number
	partNum = sheet.cell(row=num+1, column=9).value # Our Part Number
	PO  = sheet.cell(row=num+1, column=3).value # PO Number
	qty = sheet.cell(row=num+1, column=4).value

	custNum  = sheet.cell(row=num+1, column=8).value 
	orderStatus  = sheet.cell(row=num+1, column=10).value 
	description  = sheet.cell(row=num+1, column=13).value 
	shipDate = sheet.cell(row=num+1, column=5).value.date()

	### Finding and writing planner codes to file
	for key in master_PCList.keys():
		if partNum in master_PCList[key]:
			plannerCode = key
			break
		else:
			plannerCode = "n/a"

	### Finding and writing on hand amount to file
	for key in onHand.keys():
		if partNum == key:
			onHandQty = onHand[key]
			break

	### Running the daimler Check, and storing Daimler information for later
	if whatSheet == "Daimler":
		tmpSO = SO.split()[0]
		
		for key in daimler862Info.keys():
			daimlerSO = str(daimler862Info[key][0])
			ourPartNum = daimler862Info[key][1]
			daimlerShipDate = daimler862Info[key][2]
			daimlerQTY = daimler862Info[key][3]

			### Checking if Sales order exist for a row in Daimlers 862 report	
			if tmpSO == daimlerSO:
				if daimlerShipDate == shipDate:
					if partNum == ourPartNum:
						if qty == daimlerQTY:
							daimler862Info.update({key : [SO,	
							   						   	  partNum,	
							   						   	  shipDate,
							   							  daimlerQTY, PO, custNum, plannerCode, onHandQty,
														  orderStatus, description]})
							break
						else:
							daimler862Info.update({key : ["{} (try line: {})".format(daimlerSO, SO.split()[2]),	
							   						   	  partNum,	
							   						   	  shipDate,
							   							  daimlerQTY, PO, custNum, plannerCode, onHandQty,
														  orderStatus, description]})
							break
	else:
		writeToFile(plannerCode, SO, partNum, PO, onHandQty, custNum, orderStatus, description, qty, shipDate)
		row[whatSheet] += 1

### Putting Daimler in order by shipdate:
#daimler862Info = sorted(daimler862Info.items(), key=lambda x: x[3], reverse=True)
daimler862Info = sorted(daimler862Info.items(), key=lambda x: x[1][2], reverse=False)

#### Now adding in Daimler orders:
whatSheet = "Daimler"
for items in daimler862Info:
	SO = str(items[1][0])
	partNum = items[1][1]
	shipDate = items[1][2]
	qty = items[1][3]
	PO = items[1][4]

	### If the order come from JobBOSS it will have a line number (hence the -) and the information below
	if "-" in SO or "try" in SO: 
		custNum = items[1][5]
		plannerCode = items[1][6]
		onHandQty = items[1][7]
		orderStatus = items[1][8]
		description = items[1][9]

	else:
		custNum = ""
		plannerCode = ""
		onHandQty = ""
		orderStatus = ""
		description = ""

	writeToFile(plannerCode, SO, partNum, PO, onHandQty, custNum, orderStatus, description, qty, shipDate)
	row[whatSheet] += 1

### Adding headers and changing cells widths
bold = workbook.add_format({'bold': True, 'num_format': 'd mmm yyyy'})
headers = ["CUST", "Planner_Codes", "SALES ORDER", "PART NUMBER", "PO", "On Hand", "PAST DUE"]
[headers.append(items) for items in mondays]
headers.append("FUTURE")
headers.append("STATUS")
headers.append("DESCRIPTION")
for key in worksheet:
	### Heading headers
	for num, item in enumerate(headers):
		worksheet[key].write(0,num, item, bold)

	### Changing cells widths
	worksheet[key].set_column(0,0, 5)
	worksheet[key].set_column(1,1, 4)
	worksheet[key].set_column(2,2, 19)
	worksheet[key].set_column(3,3, 12)
	worksheet[key].set_column(4,out_How_Many_Weeks+7, 11)
	worksheet[key].set_column(out_How_Many_Weeks+8,out_How_Many_Weeks+8, 10.5)
	worksheet[key].set_column(out_How_Many_Weeks+9,out_How_Many_Weeks+9, 45)
	worksheet[key].freeze_panes(1, 0)
	worksheet[key].autofilter('A1:D1')
	worksheet[key].autofilter('N1')
workbook.close()
