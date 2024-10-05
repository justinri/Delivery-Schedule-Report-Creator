import openpyxl, xlsxwriter
from pathlib import Path
import datetime, string


#### Notes
"""
Seems to be skipping the last item in SH_DeliverySchedule
Run Rqmts sum needs adjusting, I think their is wrong
"""




#******************************** Only Edit **************************/
### Enter filename here:
filename = "SH_DeliveryScheduleTest"
stockFilename = "MC_StockStatus"
Output_Filename = "26wk"
out_How_Many_Weeks = 26

#******************************** STOP!!!!! **************************/

###### Finding on hand amounts
###onHand = {}
###wb_obj = openpyxl.load_workbook(stockFilename + ".xlsx") 
###sheet = wb_obj.active

###for num, rows in enumerate(sheet.iter_rows()):
###	if num == 0 or num == 1: continue    ### Skpping the header rows
###	for num2, cell in enumerate(rows):
###		if num2 != 0: break
###		try:
###			qty_onHand_units = str(sheet.cell(row=num+1, column=3).value) + " " + sheet.cell(row=num+1, column=2).value
###			onHand.update({cell.value : qty_onHand_units})
###		except TypeError:
###			pass

####*****************************************************************/
###### Creating a master list of planner codes to check against
###Planner_Codes = ["CN","CP","DC","HC","HR","HV","LD","LK","M","MF","PP","PR","TB","TC","TR","V"]
###master_PCList = {tmp:[] for tmp in Planner_Codes}

###wb_obj = openpyxl.load_workbook("MasterPartList.xlsx") 
###sheet = wb_obj.active

###for num, rows in enumerate(sheet.iter_rows()):
###	if num == 0: continue    ### Skpping the header row
###	for num2, cell in enumerate(rows):
###		if cell.value != None:
###			master_PCList[Planner_Codes[num2]].append(cell.value)

#*****************************************************************/
wb_obj = openpyxl.load_workbook(filename + ".xlsx") 
sheet = wb_obj.active

# Create a workbook and add a worksheet.
workbook = xlsxwriter.Workbook(Output_Filename + ".xlsx")
worksheet = { "MPS" : workbook.add_worksheet("MPS"), 
			  "Daimler" : workbook.add_worksheet("Daimler"),
			  "Daimler Aftermarket" : workbook.add_worksheet("Daimler Aftermarket"),
			  "Navistar Production" : workbook.add_worksheet("Navistar Production"),
			  "Navistar Service" : workbook.add_worksheet("Navistar Service"),
			  "AEES" : workbook.add_worksheet("AEES"),
			  "John Deere" : workbook.add_worksheet("John Deere"),
			  "BPC" : workbook.add_worksheet("BPC")}
whatSheet = "MPS"
alphabet_string = string.ascii_uppercase
### May have to have it create a longer string in the future, if so, have it do it automatically
columnsLetters = list(alphabet_string)
columnsLetters = columnsLetters + ["{}{}".format(columnsLetters[0],letters) for letters in columnsLetters]
row  = { "MPS" : 9, "Daimler" : 9, "Daimler Aftermarket" : 9, "Navistar Production" : 9, "Navistar Service" : 9, "AEES" : 9, "John Deere" : 9, "BPC" : 9}
col  = { "MPS" : 0, "Daimler" : 0, "Daimler Aftermarket" : 0, "Navistar Production" : 0, "Navistar Service" : 0, "AEES" : 0, "John Deere" : 0, "BPC" : 0}
multiplier = { "MPS" : 0, "Daimler" : 0, "Daimler Aftermarket" : 0, "Navistar Production" : 0, "Navistar Service" : 0, "AEES" : 0, "John Deere" : 0, "BPC" : 0}

today = datetime.date.today()
mondays = [today + datetime.timedelta(days=-today.weekday(), weeks=num) for num in range(out_How_Many_Weeks)]

### Finding months and number of weeks in the month
months = {}
for monday in mondays:
	date = monday.strftime("%b-%y")
	if date not in months:
		months.update({date : 1})
	else:
		months[date] = months[date] + 1

### Setting columns width
[worksheet[key].set_column(0,0, 15) for key in worksheet.keys()]  ### A
[worksheet[key].set_column(1,1, 50) for key in worksheet.keys()]  ### B

#*****************************************************************/
### Formatting the upper left hand corner
merge_format = workbook.add_format({
    'bold': True,
    'border': 2,
    'align': 'center',
    'valign': 'vcenter', 
	'font_size':36,
	'font_name': 'Calibri'})
dateFormat = workbook.add_format({'bold': True, 'num_format': 'd-mmm','underline':True,'align': 'center','bottom':True,'left':True,'right':True})
bold = workbook.add_format({'bold': True})
yellow = workbook.add_format({'bg_color': 'yellow'})


### Creating big header
for key in worksheet.keys():
	worksheet[key].merge_range('A1:F8', '{} 6 Month Forecast'.format(key), merge_format)
	worksheet[key].merge_range('D9:E9', 'Order Material', workbook.add_format({'bg_color': 'orange','font_size':10, 'bold':True, 'top':True,'center_across':True}))
	worksheet[key].merge_range('F9:G9', 'Need RM in House', workbook.add_format({'bg_color': 'blue','font_size':10, 'bold':True, 'top':True,'center_across':True}))
	worksheet[key].merge_range('H9:I9', 'Mfg Time', workbook.add_format({'bg_color': 'yellow','font_size':10, 'bold':True, 'top':True,'center_across':True}))
	worksheet[key].merge_range('J9:K9', 'Unmet Requirements', workbook.add_format({'bg_color': 'red','font_size':10, 'bold':True, 'top':True,'center_across':True}))
	worksheet[key].merge_range('L9:M9', 'No Requirements', workbook.add_format({'bg_color': '#c0c0c0','font_size':10, 'bold':True, 'top':True,'center_across':True}))
	worksheet[key].merge_range('N9:O9', 'Complete', workbook.add_format({'bg_color': 'green','font_size':10, 'bold':True, 'top':True,'center_across':True}))

for num, rows in enumerate(sheet.iter_rows()):
	if num == 0: continue    ### Skpping the header row

	### Does not work for some reason, unless you delete it first... but then you can undo?!?!?!?!
	### Skipping the last line of excel, it says "* Indicates a Job associated with a Sales Order."
#	if sheet.cell(row=num+1, column=1).value == "* Indicates a Job associated with a Sales Order.": continue
#	if sheet.cell(row=num+1, column=1).value[0] == "*": break

	### Checking to see what company it is
	custNum = sheet.cell(row=num+1, column=8).value
	if custNum == "2768A" or custNum == "2768B": whatSheet = "Daimler"
	elif custNum == "2768S": whatSheet = "Daimler Aftermarket"
	elif custNum == "5916A" or custNum == "5916B" or custNum == "5916L" or custNum == "5916LT" or custNum == "5916U" or custNum == "5934": whatSheet = "Navistar Production"
	elif custNum == "5916C" or custNum == "5912": whatSheet = "Navistar Service"
	elif custNum == "0160": whatSheet = "AEES"
	elif custNum == "1880" or custNum == "1881" or custNum == "1878" or custNum == "1870" or custNum == "1876" or custNum == "1879" or custNum == "100148" or custNum == "1874": whatSheet = "John Deere"
	elif custNum == "0845": whatSheet = "BPC"
	else: whatSheet = "MPS"

	### In putting static information (most of column A, dates, etc...)
	worksheet[whatSheet].write(row[whatSheet]+3+9*multiplier[whatSheet], 0, "Week of:",workbook.add_format({'bold': True,'underline':True, 'align':'right'}))
	worksheet[whatSheet].write(row[whatSheet]+3+9*multiplier[whatSheet], 1, "Qty Per",workbook.add_format({'bold': True,'underline':True, 'center_across':True}))
	[worksheet[whatSheet].write(row[whatSheet]+3+9*multiplier[whatSheet], tmpNum+2, monday, dateFormat) for tmpNum, monday in enumerate(mondays)]
	worksheet[whatSheet].write(row[whatSheet]+4+9*multiplier[whatSheet], 0, "On Hand")
	worksheet[whatSheet].write(row[whatSheet]+5+9*multiplier[whatSheet], 0, "Rqmts")
	worksheet[whatSheet].write(row[whatSheet]+6+9*multiplier[whatSheet], 0, "Delivered Qty")
	worksheet[whatSheet].write(row[whatSheet]+7+9*multiplier[whatSheet], 0, "O/H Balance")
	worksheet[whatSheet].write(row[whatSheet]+8+9*multiplier[whatSheet], 0, "Run Rqmts",yellow)
	formulaRunRqmts = "=SUM(C{}:{}{})".format(row[whatSheet]+8+9*multiplier[whatSheet],columnsLetters[-1],row[whatSheet]+8+9*multiplier[whatSheet])
	worksheet[whatSheet].write_formula("B{}".format(row[whatSheet]+9+9*multiplier[whatSheet]), formulaRunRqmts, yellow) 
	worksheet[whatSheet].write(row[whatSheet]+9+9*multiplier[whatSheet], 0, "EAU", yellow)
	worksheet[whatSheet].write(row[whatSheet]+9+9*multiplier[whatSheet], 1, 0, yellow)
	worksheet[whatSheet].write(row[whatSheet]+10+9*multiplier[whatSheet], 0, "Run Vol",yellow)
	worksheet[whatSheet].write(row[whatSheet]+10+9*multiplier[whatSheet], 1, 0,yellow)

	### Creating O/H Balance for formula
	for tmpNum2 in range(out_How_Many_Weeks):
		currentCell = columnsLetters[tmpNum2+2]+"{}".format(row[whatSheet]+8+9*multiplier[whatSheet])
		cellBefore  = columnsLetters[tmpNum2+1]+"{}".format(row[whatSheet]+8+9*multiplier[whatSheet])
		OHRow = columnsLetters[tmpNum2+2]+"{}".format(row[whatSheet]+5+9*multiplier[whatSheet])
		RqmtsRow = columnsLetters[tmpNum2+2]+"{}".format(row[whatSheet]+6+9*multiplier[whatSheet])
		DeliveredRow =columnsLetters[tmpNum2+2]+"{}".format(row[whatSheet]+7+9*multiplier[whatSheet])
		formula = '=IF({}<0,IF({}+{}=0,{}-{},({}+{})+({}-{})),IF({}+{}=0,{}-{},IF({}-{}>=0,({}+{})-({}-{}),({}+{})+({}-{}))))'.format(cellBefore,OHRow,DeliveredRow, cellBefore, RqmtsRow, OHRow, DeliveredRow, cellBefore, RqmtsRow, OHRow, DeliveredRow, cellBefore,RqmtsRow, cellBefore, RqmtsRow, OHRow, DeliveredRow, cellBefore, RqmtsRow, OHRow, DeliveredRow, cellBefore, RqmtsRow)
		worksheet[whatSheet].write_formula(currentCell, formula, bold) 

	### Creating month headers
	tmpNum = 0
	for monthKey in months.keys():
		rowMonth = row[whatSheet]+2+9*multiplier[whatSheet]
		dateRange = "{}{}:{}{}".format(columnsLetters[tmpNum+2],rowMonth+1,columnsLetters[tmpNum+1+months[monthKey]],rowMonth+1)
		worksheet[whatSheet].merge_range(dateRange, monthKey, workbook.add_format({'bold': True, 'align': 'center','top':True,'left':True,'right':True}))
		tmpNum += months[monthKey]

	### Filling out tables with all zeros to be replace with real numbers below
	for tmpNum in range(3):
		[worksheet[whatSheet].write(row[whatSheet]+4+tmpNum+9*multiplier[whatSheet], tmpNum2+2, 0.0) for tmpNum2 in range(out_How_Many_Weeks)]

	for num2, cell in enumerate(rows):
		if num2 == 8: worksheet[whatSheet].write(row[whatSheet]+9*multiplier[whatSheet], 0, cell.value,workbook.add_format({'bold': True,'center_across':True}))   ### Part number
		elif num2 == 12: worksheet[whatSheet].write(row[whatSheet]+1+9*multiplier[whatSheet], 1, cell.value,workbook.add_format({'bold': True, 'font_color':'red','center_across':True}))   ### Description

		elif num2 == 4: 
			if None == cell.value: continue
			shipDate = cell.value.date()

#			if shipDate < mondays[0]: tmpCol = 6 ### Then date is past due
#				

#				#### Figure out what to do with past dues and future
#			elif shipDate > mondays[-1]: tmpCol = out_How_Many_Weeks+7 ### Then date is a future order

			if shipDate < mondays[0]: continue
			elif shipDate > mondays[-1]: continue ### Then date is a future

#			else:
			whatMonday = shipDate + datetime.timedelta(days=-shipDate.weekday(), weeks=0)
			for num3, tmpMonday in enumerate(mondays):
				if whatMonday == tmpMonday:
					tmpCol = 2+num3
			worksheet[whatSheet].write(row[whatSheet]+5+9*multiplier[whatSheet], tmpCol, sheet.cell(row=num+1, column=4).value)


#			### Finding planner codes
#			for key in master_PCList.keys():
#				if cell.value in master_PCList[key]:
#					worksheet[whatSheet].write(row[whatSheet], 1, key)
#					break

#			else:
#				worksheet[whatSheet].write(row[whatSheet], 1, "n/a")

#			### Finding on hand amount
#			for key in onHand.keys():
#				if cell.value == key:
#					worksheet[whatSheet].write(row[whatSheet], 5, onHand[key])
#					break
#			
#		elif num2 == 2: worksheet[whatSheet].write(row[whatSheet], 4, cell.value)   ### POs
#		elif num2 == 4: 
#			if None == cell.value: continue
#			shipDate = cell.value.date()
#			if shipDate < mondays[0]: tmpCol = 6 ### Then date is past due
#			elif shipDate > mondays[-1]: tmpCol = out_How_Many_Weeks+7 ### Then date is a future order
#			else:
#				whatMonday = shipDate + datetime.timedelta(days=-shipDate.weekday(), weeks=0)
#				for num3, tmpMonday in enumerate(mondays):
#					if whatMonday == tmpMonday:
#						tmpCol = 7+num3
#			worksheet[whatSheet].write(row[whatSheet], tmpCol, sheet.cell(row=num+1, column=4).value)
#		elif num2 == 7: worksheet[whatSheet].write(row[whatSheet], 0, cell.value)  ### Customer Number
#		elif num2 == 12: worksheet[whatSheet].write(row[whatSheet], out_How_Many_Weeks+8, cell.value)
	row[whatSheet] += 1
	multiplier[whatSheet] += 1



#### Adding headers and changing cells widths
#bold = workbook.add_format({'bold': True, 'num_format': 'd mmm yyyy'})
#headers = ["CUST", "Planner_Codes", "SALES ORDER", "PART NUMBER", "PO", "On Hand", "PAST DUE"]
#[headers.append(items) for items in mondays]
#headers.append("FUTURE")
#headers.append("DESCRIPTION")
#for key in worksheet:
#	### Heading headers
#	for num, item in enumerate(headers):
#		worksheet[key].write(0,num, item, bold)

#	### Changing cells widths
#	worksheet[key].set_column(0,0, 7)
#	worksheet[key].set_column(1,2, 16)
#	worksheet[key].set_column(3,3, 16)
#	worksheet[key].set_column(4,out_How_Many_Weeks+7, 15)
#	worksheet[key].set_column(out_How_Many_Weeks+8,out_How_Many_Weeks+8, 45)
#	worksheet[key].freeze_panes(1, 0)
#	worksheet[key].autofilter('A1:D1')
workbook.close()
