########## Release Notes:
import openpyxl, xlsxwriter
from pathlib import Path
import datetime

def checkDaimler(out_How_Many_Weeks=5, filenameDaimler="862"):
	## Storing all of Daimler's ship to information
	wbDaimlerinfo = openpyxl.load_workbook('daimler_check/daimler_info.xlsx')
	sheetShipTo = wbDaimlerinfo["ship_to"]
	sheetPartNumbers = wbDaimlerinfo["DaimlerPartNumbersConverter"]

	### Getting Ship to information
	shipTo = {"A01470":{}, "A10550":{}}
	for num, rows in enumerate(sheetShipTo.iter_rows()):
		if num <= 1: continue    ### Skpping the headers row
		elif sheetShipTo.cell(row=num+1, column=2).value == None: break	

		### For PO  A01470
		SONumber = sheetShipTo.cell(row=num+1, column=1).value
		shipToLoc = sheetShipTo.cell(row=num+1, column=2).value
		shipTo["A01470"].update({shipToLoc : SONumber})

		### For PO  A10550
		SONumber = sheetShipTo.cell(row=num+1, column=4).value
		shipToLoc = sheetShipTo.cell(row=num+1, column=5).value
		shipTo["A10550"].update({shipToLoc: SONumber})

	### Getting Daimler and our part number information
	partNumbers = {}
	for num, rows in enumerate(sheetPartNumbers.iter_rows()):
		if num == 0: continue    ### Skpping the header row
		elif sheetPartNumbers.cell(row=num+1, column=2).value == None: break	

		### For PO  A01470
		daimlerPN = sheetPartNumbers.cell(row=num+1, column=1).value
		ourPN = sheetPartNumbers.cell(row=num+1, column=2).value
		partNumbers.update({daimlerPN : ourPN})

	wb_objDaimler = openpyxl.load_workbook(filenameDaimler + ".xlsx") 
	sheetDaimler = wb_objDaimler.active

	### Getting Mondays
	today = datetime.date.today()
	mondays = [today + datetime.timedelta(days=-today.weekday(), weeks=num) for num in range(out_How_Many_Weeks)]

	### Putting all daimler's 862 info a dictionary 
	daimler862Info = {}

	### Looping through Jobs
	for num, rows in enumerate(sheetDaimler.iter_rows()):
		if num == 0: continue    ### Skpping the header row
		elif sheetDaimler.cell(row=num+1, column=1).value == None: break	

		### Getting the sales order number from JobBOSS
		PO = sheetDaimler.cell(row=num+1, column=16).value # Daimler PO
		loc = sheetDaimler.cell(row=num+1, column=1).value # Daimler location
		salesOrderNum = shipTo[PO][loc]					   # Sales order number in JobBOSS

		### Getting Our Part Numer
		daimlerPartNumber = sheetDaimler.cell(row=num+1, column=2).value # Daimler Part number
		ourPartNumber = partNumbers[daimlerPartNumber]

		### Converting Daimler ship dates to our
		shipDate = sheetDaimler.cell(row=num+1, column=4).value
		shipDate = datetime.datetime.strptime(shipDate, '%Y-%m-%d').date()

		#### Information from Daimler's 862
		daimler862Info.update({"row {}".format(num) : [salesOrderNum,	
							   						   ourPartNumber,	
							   						   shipDate,	# Daimler Ship date
							   sheetDaimler.cell(row=num+1, column=5).value, # Daimler Qty
														PO]})	
	return daimler862Info

if __name__ == '__main__':
	daimler862Info = checkDaimler()


